"""
namutoxlsx.py — 나무위키 표 문법 → Excel (.xlsx) 변환기

지원:
  - || 분리 파싱 (빈 문자열 토큰 → colspan 누적)
  - ||||||  방식과 <-N> 방식 모두 처리
  - <|N> rowspan
  - 큰 colspan clamp (실제 그리드 폭 기준)
  - 텍스트 추출 (색상/링크/각주/마크업 제거)
  - [br] → 엑셀 개행
  - 병합 셀 openpyxl 적용

사용:
  python namutoxlsx.py input.txt output.xlsx
  python namutoxlsx.py input.txt output.xlsx --sheet "시간표"
  cat input.txt | python namutoxlsx.py - output.xlsx
"""

import re
import sys
import argparse
import warnings
import textwrap

try:
    import openpyxl
    from openpyxl.styles import Alignment, PatternFill
except ImportError:
    print(textwrap.dedent("""
        [오류] openpyxl 모듈이 없습니다. 아래 방법으로 설치하세요.

        일반 pip:
          pip install openpyxl

        가상환경(venv/conda):
          (venv) pip install openpyxl
          conda install openpyxl

        시스템 Python(Debian/Ubuntu, pip 제한 환경):
          pip install --break-system-packages openpyxl
          또는: sudo apt install python3-openpyxl

        Windows(py 런처):
          py -m pip install openpyxl
    """).strip(), file=sys.stderr)
    sys.exit(1)

# openpyxl 경고 억제 (UserWarning: Unknown extension 등)
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


# ──────────────────────────────────────────────
# 1. 텍스트 정제
# ──────────────────────────────────────────────

def strip_markup(text):
    """나무위키 마크업을 제거하고 순수 텍스트만 반환."""
    if not text:
        return ""

    # [[파일:...]] 제거
    text = re.sub(r'\[\[파일:[^\]]*\]\]', '', text)

    # [*각주] 제거 — 내부 [[링크]] 보호 후 제거
    def remove_footnotes(s):
        protected = []
        def protect(m):
            protected.append(m.group(0))
            return f'\x00{len(protected)-1}\x00'
        s = re.sub(r'\[\[[^\]]*\]\]', protect, s)
        s = re.sub(r'\[\*[^\]]*\]', '', s)
        for i, orig in enumerate(protected):
            s = s.replace(f'\x00{i}\x00', orig)
        return s
    text = remove_footnotes(text)

    # {{{ }}} 반복 처리 (중첩 대응)
    for _ in range(5):
        prev = text
        text = re.sub(r'\{\{\{#[\w,#]+\s(.*?)\}\}\}', r'\1', text, flags=re.DOTALL)
        text = re.sub(r'\{\{\{[+-]\d\s(.*?)\}\}\}', r'\1', text, flags=re.DOTALL)
        text = re.sub(r'\{\{\{(.*?)\}\}\}', r'\1', text, flags=re.DOTALL)
        if text == prev:
            break

    # 이스케이프된 대괄호를 placeholder로 보호 (링크 파싱망 회피)
    text = re.sub(r'\\\[', '\x01', text)
    text = re.sub(r'\\\]', '\x02', text)
    # 나머지 이스케이프 문법 문자 해제 (경로 등 일반 백슬래시는 보존)
    text = re.sub(r'\\([\\{|}\'~\-_*#])', r'\1', text)

    # [br] → 개행 (링크 처리 전에)
    text = text.replace('[br]', '\n')

    # [[링크|표시텍스트]] → 표시텍스트 (개행 포함 DOTALL)
    text = re.sub(r'\[\[([^\]|]+)\|(.+?)\]\]', r'\2', text, flags=re.DOTALL)
    # [[링크]] → 링크
    text = re.sub(r'\[\[([^\]]+)\]\]', r'\1', text)

    text = re.sub(r'\[include\([^\)]*\)\]', '', text)
    text = re.sub(r'\[youtube\([^\)]*\)\]', '', text)
    text = text.replace("'''", '').replace("''", '')
    text = text.replace('__', '').replace('~~', '')
    # placeholder 복원
    text = text.replace('\x01', '[').replace('\x02', ']')
    return text.strip()


# ──────────────────────────────────────────────
# 2. 행 파싱
# ──────────────────────────────────────────────

def parse_cell_attrs(attr_str):
    """<-N>, <|N> 속성 문자열에서 colspan/rowspan 추출."""
    colspan = 1
    rowspan = 1
    m = re.search(r'<-(\d+)>', attr_str)
    if m:
        colspan = int(m.group(1))
    m = re.search(r'<\|(\d+)>', attr_str)
    if m:
        rowspan = int(m.group(1))
    return colspan, rowspan


def split_row(line):
    """
    나무위키 표 한 행 → 셀 토큰 리스트.
    반환: list of (attr_str, content_str)

    규칙:
      빈 토큰("") → 앞 셀 colspan +1  (||||||  방식)
      공백 포함 토큰(" ") → 별개의 빈 셀
    """
    line = line.strip()
    if not line.startswith('||'):
        return []
    if line.endswith('||'):
        line = line[:-2]
    line = line[2:]

    raw_tokens = line.split('||')
    cells = []
    pending = 0  # 빈 토큰 누적

    for tok in raw_tokens:
        if tok == '':
            pending += 1
        else:
            if pending > 0 and cells:
                attr, content = cells[-1]
                m = re.search(r'<-(\d+)>', attr)
                if m:
                    old = int(m.group(1))
                    attr = re.sub(r'<-\d+>', f'<-{old + pending}>', attr)
                else:
                    attr = f'<-{1 + pending}>' + attr
                cells[-1] = (attr, content)
                pending = 0

            attr_match = re.match(r'^((?:<[^>]*>)*)(.*)', tok, re.DOTALL)
            if attr_match:
                attr_str = attr_match.group(1)
                content_str = attr_match.group(2)
            else:
                attr_str = ''
                content_str = tok
            cells.append((attr_str, content_str))

    # 루프 끝 잔여 빈 토큰
    if pending > 0 and cells:
        attr, content = cells[-1]
        m = re.search(r'<-(\d+)>', attr)
        if m:
            old = int(m.group(1))
            attr = re.sub(r'<-\d+>', f'<-{old + pending}>', attr)
        else:
            attr = f'<-{1 + pending}>' + attr
        cells[-1] = (attr, content)

    return cells


# ──────────────────────────────────────────────
# 3. 그리드 구성
# ──────────────────────────────────────────────

def extract_table_lines(text):
    return [l.strip() for l in text.splitlines() if l.strip().startswith('||')]


def build_grid(table_lines):
    parsed_rows = [split_row(l) for l in table_lines]
    if not parsed_rows:
        return [], 0, 0

    # 1패스: 그리드 최대 폭 추정
    max_cols = 0
    for cells in parsed_rows:
        col_sum = sum(parse_cell_attrs(attr)[0] for attr, _ in cells)
        if col_sum > max_cols:
            max_cols = col_sum

    num_rows = len(parsed_rows)
    PAD = 50
    grid = [[None] * (max_cols + PAD) for _ in range(num_rows + PAD)]
    occupied = {}

    for r, cells in enumerate(parsed_rows):
        c = 0
        for attr_str, content_str in cells:
            while occupied.get((r, c)):
                c += 1

            colspan, rowspan = parse_cell_attrs(attr_str)

            # colspan clamp
            remaining = max_cols - c
            if remaining < 1:
                remaining = 1
            colspan = min(colspan, remaining)
            if colspan < 1:
                colspan = 1

            text = strip_markup(content_str)
            bgcolor = None
            m = re.search(r'<bgcolor=#?([0-9a-fA-F]{3,6})', attr_str)
            if m:
                raw = m.group(1).upper()
                # 3자리 → 6자리 확장 (f00 → ff0000)
                bgcolor = ''.join(c*2 for c in raw) if len(raw) == 3 else raw
            grid[r][c] = {'text': text, 'colspan': colspan, 'rowspan': rowspan, 'bgcolor': bgcolor}

            for dr in range(rowspan):
                for dc in range(colspan):
                    if dr == 0 and dc == 0:
                        continue
                    occupied[(r + dr, c + dc)] = True

            c += colspan

    return grid, num_rows, max_cols


# ──────────────────────────────────────────────
# 4. 엑셀 출력
# ──────────────────────────────────────────────

def write_xlsx(grid, num_rows, num_cols, output_path, sheet_name='Sheet1'):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    for r in range(num_rows):
        for c in range(num_cols):
            cell_data = grid[r][c]
            if cell_data is None:
                continue

            text = cell_data['text']
            colspan = cell_data['colspan']
            rowspan = cell_data['rowspan']
            bgcolor = cell_data.get('bgcolor')

            er, ec = r + 1, c + 1
            cell = ws.cell(row=er, column=ec, value=text)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            if bgcolor:
                cell.fill = PatternFill(start_color='FF' + bgcolor,
                                        end_color='FF' + bgcolor,
                                        fill_type='solid')

            if colspan > 1 or rowspan > 1:
                ws.merge_cells(
                    start_row=er, start_column=ec,
                    end_row=er + rowspan - 1,
                    end_column=ec + colspan - 1
                )

    wb.save(output_path)
    print(f"저장 완료: {output_path}  ({num_rows}행 × {num_cols}열)", file=sys.stderr)


# ──────────────────────────────────────────────
# 5. 메인
# ──────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        prog='namutoxlsx',
        description='나무위키 표 문법 → Excel(.xlsx) 변환기',
        epilog='예시:\n  python namutoxlsx.py input.txt output.xlsx\n  cat raw.txt | python namutoxlsx.py - out.xlsx',
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument('input', help='입력 파일 경로 (- 이면 stdin)')
    parser.add_argument('output', help='출력 .xlsx 파일 경로')
    parser.add_argument('--sheet', default='Sheet1', help='시트 이름 (기본: Sheet1)')
    args = parser.parse_args()

    if args.input == '-':
        text = sys.stdin.read()
    else:
        with open(args.input, encoding='utf-8') as f:
            text = f.read()

    table_lines = extract_table_lines(text)
    if not table_lines:
        print("[오류] 표 행(|| 시작)을 찾지 못했습니다.", file=sys.stderr)
        sys.exit(1)

    print(f"표 행 {len(table_lines)}줄 발견", file=sys.stderr)
    grid, num_rows, num_cols = build_grid(table_lines)
    write_xlsx(grid, num_rows, num_cols, args.output, sheet_name=args.sheet)


if __name__ == '__main__':
    main()
