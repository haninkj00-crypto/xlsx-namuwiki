import os
import argparse
import datetime
import warnings

try:
    import openpyxl
    from openpyxl.utils import range_boundaries
except ImportError:
    import sys, textwrap
    print(textwrap.dedent("""
        [오류] openpyxl 모듈이 없습니다. 아래 방법으로 설치하세요.

        일반 pip:
          pip install openpyxl

        가상환경(venv/conda):
          (venv) pip install openpyxl
          conda install openpyxl

        시스템 Python(Debian/Ubuntu, pip 제한 환경):
          pip install --break-system-packages openpyxl
          또는: sudo apt install python3-openpyxl

        Windows(py 런처):
          py -m pip install openpyxl
    """).strip(), file=sys.stderr)
    sys.exit(1)

# openpyxl 경고 억제 (UserWarning: Unknown extension 등)
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

def escape_namuwiki_syntax(text, apply_options=None):
    """
    텍스트 내의 나무위키 마크업을 무효화합니다. 
    폰트 서식 관련 마크업은 apply_options와 무관하게 항상 강제 무효화됩니다.
    """
    if not text:
        return ""
    if apply_options is None:
        apply_options = []

    text = text.replace("\\", "\\\\")
    
    text = text.replace("'''", "\\'\\'\\'")
    text = text.replace("''", "\\'\\'")
    text = text.replace("__", "\\_\\_")
    text = text.replace("~~", "\\~\\~")
    text = text.replace("--", "\\-\\-")
    
    token_map = {
        'intlink': ["[[", "]]"],
        'extlink': ["[http", "]"],
        'footnote': ["[*"],
        'macro': ["{{{#", "[include", "[youtube"],
        'table': ["||"]
    }
    
    for option, tokens in token_map.items():
        if option not in apply_options:
            for token in tokens:
                text = text.replace(token, "\\" + token)
                
    return text

def apply_excel_font_style(cell, text):
    """
    openpyxl의 셀 폰트 속성을 분석하여 나무위키 마크업으로 텍스트를 감쌉니다.
    """
    if getattr(cell, 'font', None) is None or not text:
        return text
        
    styled_text = text
    
    if cell.font.bold:
        styled_text = f"'''{styled_text}'''"
    if cell.font.italic:
        styled_text = f"''{styled_text}''"
    if cell.font.underline:
        styled_text = f"__{styled_text}__"
    if cell.font.strike:
        styled_text = f"~~{styled_text}~~"
        
    return styled_text

def excel_to_namuwiki(file_path, cell_range=None, sheet_name=None, apply_options=None, fx_mode='output'):
    """
    Excel 데이터를 읽고 폰트 속성, 서식, 병합 정보를 반영하여 나무위키 표로 변환합니다.
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {file_path}")
    
    is_data_only = (fx_mode == 'output')
    wb = openpyxl.load_workbook(file_path, data_only=is_data_only)
    ws = wb[sheet_name] if sheet_name else wb.worksheets[0]

    if cell_range:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    else:
        min_col, min_row = 1, 1
        max_col, max_row = ws.max_column, ws.max_row

    # 사전 작업: 병합된 셀 좌표 맵 생성
    merge_map = {}
    for merged_range in ws.merged_cells.ranges:
        m_min_col, m_min_row, m_max_col, m_max_row = range_boundaries(str(merged_range))
        for r in range(m_min_row, m_max_row + 1):
            for c in range(m_min_col, m_max_col + 1):
                merge_map[(r, c)] = {
                    'is_top_left': (r == m_min_row and c == m_min_col),
                    'rowspan': m_max_row - m_min_row + 1,
                    'colspan': m_max_col - m_min_col + 1
                }

    # 표 시작 마크업
    namuwiki_lines = ['{{{#!wiki style="word-break: keep-all"']
    
    # ---------------------------------------------------------
    # 메인 루프 시작
    # ---------------------------------------------------------
    for r in range(min_row, max_row + 1):
        row_tokens = []
        is_row_empty = True
        
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            
            # 1. 엑셀 셀 원본 텍스트 추출 및 서식 검증
            if cell.value is not None:
                if cell.number_format not in ['General', '@'] and not cell.is_date:
                    if isinstance(cell.value, (int, float)):
                        raw_val = str(format_number(cell.value, cell.number_format))
                    else:
                        raise ValueError(f"서식 오류 중단: {cell.coordinate} 셀의 서식을 확인하십시오.")
                elif fx_mode == 'output' and cell.is_date:
                    if isinstance(cell.value, datetime.datetime):
                        if cell.value.time() == datetime.time(0, 0):
                            raw_val = cell.value.strftime('%Y-%m-%d')
                        else:
                            raw_val = cell.value.strftime('%Y-%m-%d %H:%M:%S')
                    elif isinstance(cell.value, datetime.time):
                        raw_val = cell.value.strftime('%H:%M:%S')
                    else:
                        raw_val = str(cell.value)
                else:
                    raw_val = str(cell.value)
                    if fx_mode == 'input' and getattr(cell, 'quotePrefix', False):
                        raw_val = "'" + raw_val
            else:
                raw_val = ""
            
            # 2. 나무위키 문법 강제 이스케이프 적용
            escaped_text = escape_namuwiki_syntax(raw_val, apply_options)
            
            # 3. 엑셀 폰트 서식 감지 및 마크업 조립 (배경색 속성은 분리)
            display_text = apply_excel_font_style(cell, escaped_text)
            
            # 4. 엑셀 줄바꿈(Alt+Enter) 치환
            display_text = display_text.replace("\n", "[br]")

            # 셀 배경색(bgcolor) 추출            
            cell_attr = ""
            if cell.fill and cell.fill.fill_type == 'solid':
                color_obj = cell.fill.start_color
                # 색상 객체가 존재하고, 타입이 rgb이며, 값이 있는 경우만 처리
                if color_obj and getattr(color_obj, 'type', None) == 'rgb' and color_obj.rgb:
                    raw_rgb = str(color_obj.rgb)
                    if raw_rgb not in ['00000000', 'FFFFFFFF', 'System Foreground']:
                        bg_hex = raw_rgb[2:].lower()
                        cell_attr += f"<bgcolor=#{bg_hex}>"

            
            # 5. 병합 셀 / 일반 셀 판단 및 나무위키 표 문법 조립
            if (r, c) in merge_map:
                info = merge_map[(r, c)]
                if not info['is_top_left']:
                    continue
                
                if info['colspan'] > 1:
                    cell_attr += f"<-{info['colspan']}>"
                if info['rowspan'] > 1:
                    cell_attr += f"<|{info['rowspan']}>"
                
                row_tokens.append(f"||{cell_attr} {display_text} ")
                is_row_empty = False
            else:
                row_tokens.append(f"||{cell_attr} {display_text} ")
                is_row_empty = False
                
        # 한 줄 작업 완료 시 추가
        if not is_row_empty and row_tokens:
            namuwiki_lines.append("".join(row_tokens) + "||")
    # ---------------------------------------------------------
    # 메인 루프 종료
    # ---------------------------------------------------------

    namuwiki_lines.append('}}}')
    return "\n".join(namuwiki_lines)

if __name__ == "__main__":
    import sys
    if len(sys.argv) == 1 or (len(sys.argv) == 2 and sys.argv[1].lower() == 'help'):
        sys.argv = [sys.argv[0], '--help']

    parser = argparse.ArgumentParser(
        prog='xlxstonamu',
        description='Excel(.xlsx) 표를 나무위키 표 문법으로 변환합니다.',
        epilog='예시:\n  xlxstonamu.py table.xlsx\n  xlxstonamu.py table.xlsx -r A1:D10 -s Sheet2\n  xlxstonamu.py table.xlsx -a intlink extlink --fx input',
        formatter_class=argparse.RawDescriptionHelpFormatter
    )

    parser.add_argument("file", help="변환할 Excel 파일 경로 (.xlsx)")
    parser.add_argument("-r", "--range", help="변환할 셀 범위 (예: A1:D5). 생략 시 시트 전체", default=None)
    parser.add_argument("-s", "--sheet", help="대상 시트명. 생략 시 첫 번째 시트", default=None)
    parser.add_argument(
        "-a", "--apply",
        nargs='+',
        choices=['intlink', 'extlink', 'footnote', 'macro', 'table', 'all'],
        help="이스케이프 없이 원형 유지할 나무위키 문법 (intlink: [[링크]], extlink: [URL], footnote: [*주석], macro: {{{매크로}}} 등, table: ||표||, all: 전체)",
        default=[]
    )
    parser.add_argument(
        "--fx",
        choices=['input', 'output'],
        default='output',
        help="수식 처리 방식. output: 계산된 결과값(기본값), input: 수식 문자열 그대로"
    )
    args = parser.parse_args()
    
    if 'all' in args.apply:
        args.apply = ['intlink', 'extlink', 'footnote', 'macro', 'table']
        
    try:
        result = excel_to_namuwiki(
            file_path=args.file,
            cell_range=args.range,
            sheet_name=args.sheet,
            apply_options=args.apply,
            fx_mode=args.fx
        )
        print(result)
    except Exception as e:
        print(f"오류 발생: {e}")